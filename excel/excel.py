import openpyxl
import json
from datetime import datetime

workbook = openpyxl.load_workbook("HORARIOS DE INGRESOS 2024.xlsx")
sheet = workbook["JUNIO 2024 "]  # hoja

data_list = []

# Definir las filas y columnas
start_row_id_name = 71  # Fila de inicio para IDs y nombres
start_col_id_name = 1  # Columna de inicio para IDs y nombres
end_row_id_name = 88  # Fila final para IDs y nombres

start_row_times = 93  # Fila de inicio para entradas y salidas
# end_row_times = 44  # Fila final para entradas y salidas

start_col_times = 13 # Columna de inicio para entradas y salidas se suma 5
end_col_times = 16# Columna final para entradas y salidas se suma 5

# Dia
timeDay = "2024-06-26"


# Formato fecha
def format_datetime(date_str, time_str):
    try:
        datetime.strptime(time_str, '%H:%M:%S')
        return f"{date_str}T{time_str}.000-05:00"
    except ValueError:
        return None


# Iterar sobre las filas y leer los IDs y nombres
print(timeDay)

for index, row in enumerate(
        sheet.iter_rows(min_row=start_row_id_name, min_col=start_col_id_name, max_col=start_col_id_name + 1,
                        values_only=True, max_row=end_row_id_name + 1)):

    # print(row)

    empleado_id = str(row[0]).strip() if row[0] else None
    nombre = str(row[1]).strip() if row[1] else None

    print(f"Empleado ID: {empleado_id}, Nombre: {nombre}")

    # Filtrar empleados sin ID o sin nombre
    if not empleado_id or empleado_id.lower() == 'none' or not nombre or nombre.lower() == 'none':
        continue

    # Iterar sobre las filas y leer los datos de entradas y salidas
    time_row = sheet.iter_rows(min_row=start_row_times + index, min_col=start_col_times, max_col=end_col_times,
                               values_only=True)

    time_row = list(time_row)[0]  # Convertir a lista y tomar la primera (y única) fila
    print('time_row')
    print(time_row)
    if len(time_row) == 4:
        entradas_salidas = [str(cell).strip() for cell in time_row]

        # Crear objetos para cada entrada y salida
        print('entradas_salidas')
        print(entradas_salidas)

        times = ["entrada1", "salida1", "entrada2", "salida2"]
        for i, time in enumerate(times):
            formatted_time = format_datetime(timeDay, entradas_salidas[i])
            if formatted_time:
                obj = {
                    "empleado": int(empleado_id),
                    # "nombre": nombre,
                    "hora": formatted_time,
                    # "rectificar": True
                }
                data_list.append(obj)
            else:

                print(f"Tiempo no válido encontrado para el empleado {empleado_id}, {nombre}: {entradas_salidas[i]}")

json_data = json.dumps(data_list, indent=2)
output_file = "../horarios_ingresos.json"
with open(output_file, 'w') as f:
    json.dump(data_list, f, indent=2)

print(json_data)
